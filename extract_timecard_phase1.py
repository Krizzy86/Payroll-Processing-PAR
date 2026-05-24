from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_FILE_NAME = "Employee Timecard.xlsx"
CONFIG_FILE_NAME = "timesheet_config.xlsx"
OUTPUT_FILE_NAME = "Phase 1 Payroll Extract.xlsx"


@dataclass
class EmployeeRecord:
    name: str
    base_pay: float | int | None
    total_hours: float | int | None


def normalize_name(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def create_default_config(config_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Config"

    sections = [
        ("Salary Employees", ["Leon, Tina"]),
        ("Non Tip Pool Employees", ["Hernandez, Aydelin"]),
        ("Ignore These Employees", []),
    ]

    row = 1
    for title, names in sections:
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="4F81BD")
        row += 1
        if names:
            for name in names:
                ws.cell(row=row, column=1, value=name)
                row += 1
        else:
            ws.cell(row=row, column=1, value="")
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 34
    wb.save(config_path)


def read_config(config_path: Path) -> dict[str, set[str]]:
    if not config_path.exists():
        create_default_config(config_path)

    wb = load_workbook(config_path, data_only=True)
    ws = wb.active

    expected_sections = {
        "salary employees": "salary",
        "non tip pool employees": "non_tip_pool",
        "ignore these employees": "ignore",
    }
    result = {"salary": set(), "non_tip_pool": set(), "ignore": set()}
    current: str | None = None

    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        value = row[0]
        if value is None:
            continue
        text = str(value).strip()
        section = expected_sections.get(text.casefold())
        if section:
            current = section
            continue
        if current and text:
            result[current].add(normalize_name(text))

    return result


def header_map(ws, row_num: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[row_num]:
        if cell.value is not None:
            headers[str(cell.value).strip().casefold()] = cell.column
    return headers


def extract_records(input_path: Path) -> list[EmployeeRecord]:
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active
    records: list[EmployeeRecord] = []
    row = 1

    while row <= ws.max_row:
        if str(ws.cell(row=row, column=1).value or "").strip().casefold() == "employe name":
            name = str(ws.cell(row=row + 2, column=2).value or "").strip()
            header_row = None
            for candidate in range(row + 1, min(row + 8, ws.max_row) + 1):
                if str(ws.cell(row=candidate, column=1).value or "").strip().casefold() == "date":
                    header_row = candidate
                    break

            if not name or header_row is None:
                row += 1
                continue

            headers = header_map(ws, header_row)
            rate_col = headers.get("rate")
            total_hours_col = headers.get("total hrs")
            base_pay = None
            total_hours = None
            block_end = header_row

            for detail_row in range(header_row + 1, ws.max_row + 1):
                first_cell = str(ws.cell(row=detail_row, column=1).value or "").strip()
                if first_cell.casefold() == "signature":
                    block_end = detail_row
                    break
                if first_cell.casefold() == "totals:":
                    if total_hours_col:
                        total_hours = ws.cell(row=detail_row, column=total_hours_col).value
                    block_end = detail_row
                    break
                if rate_col and base_pay is None:
                    rate = ws.cell(row=detail_row, column=rate_col).value
                    if rate not in (None, ""):
                        base_pay = rate

            records.append(EmployeeRecord(name=name, base_pay=base_pay, total_hours=total_hours))
            row = block_end + 1
            continue

        row += 1

    return records


def sort_records(records: list[EmployeeRecord], config: dict[str, set[str]]) -> list[EmployeeRecord]:
    ignored = config["ignore"]

    def group(record: EmployeeRecord) -> int:
        key = normalize_name(record.name)
        if key in config["salary"]:
            return 0
        if key in config["non_tip_pool"]:
            return 1
        return 2

    visible_records = [record for record in records if normalize_name(record.name) not in ignored]
    return sorted(visible_records, key=lambda record: (group(record), normalize_name(record.name)))


def write_output(records: list[EmployeeRecord], output_path: Path, config: dict[str, set[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Phase 1 Extract"
    headers = ["Name", "Total Hours", "Base Hourly", "Guaranteed Rate"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")

    for record in records:
        is_non_tip_pool = normalize_name(record.name) in config["non_tip_pool"]
        guaranteed_rate = record.base_pay if is_non_tip_pool else (
            record.base_pay + 3 if record.base_pay and record.base_pay > 0 else 0
        )
        ws.append([record.name, record.total_hours, record.base_pay, guaranteed_rate])

    widths = [32, 12, 14, 18]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    for cell in ws.iter_cols(min_row=2, min_col=2, max_col=2):
        for item in cell:
            item.number_format = "0.00"

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = "$#,##0.00"

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Create phase-one payroll extract from Employee Timecard.xlsx.")
    parser.add_argument("pay_period_folder", type=Path)
    parser.add_argument("--config", type=Path, default=None)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()

    pay_period_folder = args.pay_period_folder
    input_path = pay_period_folder / INPUT_FILE_NAME
    config_path = args.config or pay_period_folder / CONFIG_FILE_NAME
    output_path = args.output or pay_period_folder / OUTPUT_FILE_NAME

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    config = read_config(config_path)
    records = sort_records(extract_records(input_path), config)
    write_output(records, output_path, config)

    print(f"Config: {config_path}")
    print(f"Output: {output_path}")
    print(f"Rows: {len(records)}")


if __name__ == "__main__":
    main()
